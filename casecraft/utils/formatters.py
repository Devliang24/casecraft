"""Output formatters for test cases."""

import json
from abc import ABC, abstractmethod
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from casecraft.config.template_manager import TemplateManager
from casecraft.models.test_case import TestCaseCollection


class OutputFormatter(ABC):
    """Abstract base class for output formatters."""
    
    @abstractmethod
    def format(self, collection: TestCaseCollection) -> Union[str, bytes]:
        """Format test case collection to string or bytes.
        
        Args:
            collection: Test case collection to format
            
        Returns:
            Formatted string or bytes
        """
        pass
    
    @abstractmethod
    def get_file_extension(self) -> str:
        """Get file extension for this format.
        
        Returns:
            File extension (including dot)
        """
        pass
    
    def is_binary(self) -> bool:
        """Check if this formatter produces binary output.
        
        Returns:
            True if binary, False if text
        """
        return False


class JSONFormatter(OutputFormatter):
    """JSON output formatter."""
    
    def __init__(self, indent: int = 2):
        """Initialize JSON formatter.
        
        Args:
            indent: JSON indentation level
        """
        self.indent = indent
    
    def format(self, collection: TestCaseCollection) -> str:
        """Format collection as JSON."""
        return collection.model_dump_json(indent=self.indent, exclude_none=True)
    
    def get_file_extension(self) -> str:
        """Get JSON file extension."""
        return ".json"


class CompactJSONFormatter(OutputFormatter):
    """Compact JSON output formatter (no indentation)."""
    
    def format(self, collection: TestCaseCollection) -> str:
        """Format collection as compact JSON."""
        return collection.model_dump_json(exclude_none=True)
    
    def get_file_extension(self) -> str:
        """Get JSON file extension."""
        return ".json"


class PrettyJSONFormatter(OutputFormatter):
    """Pretty-printed JSON formatter with extra formatting."""
    
    def format(self, collection: TestCaseCollection) -> str:
        """Format collection as pretty JSON with metadata."""
        data = collection.model_dump(exclude_none=True)
        
        # Add formatting metadata
        formatted_data = {
            "_metadata": {
                "format": "CaseCraft Test Cases",
                "version": "1.0",
                "endpoint": f"{collection.method} {collection.path}",
                "total_test_cases": len(collection.test_cases)
            },
            **data
        }
        
        return json.dumps(formatted_data, indent=2, ensure_ascii=False, sort_keys=False)
    
    def get_file_extension(self) -> str:
        """Get JSON file extension."""
        return ".json"


class ExcelFormatter(OutputFormatter):
    """Excel (XLSX) output formatter."""
    
    def __init__(self, template_manager: Optional[TemplateManager] = None):
        """Initialize Excel formatter.
        
        Args:
            template_manager: Template configuration manager
        """
        self.template_manager = template_manager or TemplateManager()
        self.columns = self.template_manager.get_excel_columns()
        self.styles = self.template_manager.get_excel_styles()
    
    def format(self, collection: TestCaseCollection) -> bytes:
        """Format collection as Excel file.
        
        Args:
            collection: Test case collection
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self._get_sheet_name(collection)
        
        # Write headers
        self._write_headers(ws)
        
        # Write test cases
        for row_idx, test_case in enumerate(collection.test_cases, start=2):
            self._write_test_case(ws, row_idx, test_case, collection)
        
        # Apply styles
        self._apply_styles(ws)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    def get_file_extension(self) -> str:
        """Get Excel file extension."""
        return ".xlsx"
    
    def is_binary(self) -> bool:
        """Excel files are binary."""
        return True
    
    def _get_sheet_name(self, collection: TestCaseCollection) -> str:
        """Generate sheet name from collection."""
        # Use method and path, limited to 31 chars (Excel limit)
        name = f"{collection.method}_{collection.path.replace('/', '_')}"
        return name[:31] if len(name) > 31 else name
    
    def _write_headers(self, ws):
        """Write header row."""
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(
            start_color=self.styles.get('header_bg_color', 'D3D3D3'),
            end_color=self.styles.get('header_bg_color', 'D3D3D3'),
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, column_config in enumerate(self.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_config['header'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _write_test_case(self, ws, row_idx, test_case, collection):
        """Write a single test case row."""
        # Map field names to values
        data_map = {
            'case_id': getattr(test_case, 'case_id', ''),
            'module': getattr(test_case, 'module', ''),
            'api_name': f"{collection.method} {collection.path}",
            'name': test_case.name,
            'priority': getattr(test_case, 'priority', ''),
            'preconditions': self._format_list(getattr(test_case, 'preconditions', [])),
            'method': test_case.method,
            'path': test_case.path,
            'headers': self._format_dict(test_case.headers),
            'request_data': self._format_request_data(test_case),
            'status': test_case.status,
            'response_assertions': self._format_response_assertions(test_case),
            'postconditions': self._format_list(getattr(test_case, 'postconditions', [])),
            'remarks': getattr(test_case, 'remarks', '')
        }
        
        # Write data and apply priority coloring
        for col_idx, column_config in enumerate(self.columns, start=1):
            field = column_config['field']
            value = data_map.get(field, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply text wrapping
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Apply priority coloring
            if field == 'priority' and value:
                priority_colors = self.styles.get('priority_colors', {})
                if value in priority_colors:
                    cell.fill = PatternFill(
                        start_color=priority_colors[value],
                        end_color=priority_colors[value],
                        fill_type="solid"
                    )
    
    def _apply_styles(self, ws):
        """Apply general styles to worksheet."""
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for col_idx, column_config in enumerate(self.columns, start=1):
            column_letter = get_column_letter(col_idx)
            
            # Use configured width or calculate based on content
            if 'width' in column_config:
                ws.column_dimensions[column_letter].width = column_config['width']
            else:
                # Calculate max width from content
                max_length = len(column_config['header'])
                for row in ws.iter_rows(min_row=2, max_col=col_idx, min_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                # Set width with a reasonable maximum
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _format_dict(self, data: Dict[str, Any]) -> str:
        """Format dictionary as readable string."""
        if not data:
            return ""
        
        lines = []
        for key, value in data.items():
            lines.append(f"{key}: {value}")
        return "\n".join(lines)
    
    def _format_list(self, data: List[str]) -> str:
        """Format list as numbered string."""
        if not data:
            return ""
        
        return "\n".join(f"{i+1}. {item}" for i, item in enumerate(data))
    
    def _format_request_data(self, test_case) -> str:
        """Format request data (params and body)."""
        parts = []
        
        if hasattr(test_case, 'path_params') and test_case.path_params:
            parts.append(f"Path参数:\n{self._format_dict(test_case.path_params)}")
        
        if hasattr(test_case, 'query_params') and test_case.query_params:
            parts.append(f"Query参数:\n{self._format_dict(test_case.query_params)}")
        
        if hasattr(test_case, 'body') and test_case.body:
            parts.append(f"请求体:\n{json.dumps(test_case.body, indent=2, ensure_ascii=False)}")
        
        return "\n\n".join(parts) if parts else ""
    
    def _format_response_assertions(self, test_case) -> str:
        """Format response assertions."""
        parts = []
        
        if hasattr(test_case, 'resp_content') and test_case.resp_content:
            parts.append(f"响应内容验证:\n{json.dumps(test_case.resp_content, indent=2, ensure_ascii=False)}")
        
        if hasattr(test_case, 'resp_schema') and test_case.resp_schema:
            # Simplify schema for display
            schema_summary = {
                "type": test_case.resp_schema.get("type", "object"),
                "properties": list(test_case.resp_schema.get("properties", {}).keys())[:5]
            }
            parts.append(f"响应结构:\n{json.dumps(schema_summary, indent=2, ensure_ascii=False)}")
        
        if hasattr(test_case, 'rules') and test_case.rules:
            parts.append(f"业务规则:\n{self._format_list(test_case.rules)}")
        
        return "\n\n".join(parts) if parts else ""


def get_formatter(format_type: str = "json", template_manager: Optional[TemplateManager] = None) -> OutputFormatter:
    """Get formatter instance by type.
    
    Args:
        format_type: Format type (json, compact, pretty, excel)
        template_manager: Optional template manager for Excel formatter
        
    Returns:
        Formatter instance
        
    Raises:
        ValueError: If format type is unsupported
    """
    formatters = {
        "json": JSONFormatter,
        "compact": CompactJSONFormatter,
        "pretty": PrettyJSONFormatter,
        "excel": ExcelFormatter
    }
    
    if format_type not in formatters:
        supported = ", ".join(formatters.keys())
        raise ValueError(f"Unsupported format type: {format_type}. Supported: {supported}")
    
    # Excel formatter needs template manager
    if format_type == "excel":
        return ExcelFormatter(template_manager)
    
    return formatters[format_type]()