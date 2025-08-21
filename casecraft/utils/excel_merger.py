"""Excel merger for combining multiple test case collections into one file."""

from io import BytesIO
from typing import List, Optional, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from casecraft.config.template_manager import TemplateManager
from casecraft.models.test_case import TestCaseCollection
from casecraft.utils.formatters import ExcelFormatter


class ExcelMerger:
    """Merges multiple TestCaseCollections into a single Excel file."""
    
    def __init__(self, template_manager: Optional[TemplateManager] = None):
        """Initialize Excel merger.
        
        Args:
            template_manager: Template configuration manager
        """
        self.template_manager = template_manager or TemplateManager()
        self.workbook = Workbook()
        self.sheets_created = 0
        
        # Track module collections for merging
        self.module_collections: Dict[str, List[TestCaseCollection]] = {}
    
    def add_collection(self, collection: TestCaseCollection, sheet_name: Optional[str] = None):
        """Add a collection to the workbook as a new sheet.
        
        Args:
            collection: Test case collection to add
            sheet_name: Optional custom sheet name
        """
        # Create or get worksheet
        if self.sheets_created == 0:
            ws = self.workbook.active
        else:
            ws = self.workbook.create_sheet()
        
        # Set sheet name
        if sheet_name:
            ws.title = self._sanitize_sheet_name(sheet_name)
        else:
            ws.title = self._generate_sheet_name(collection)
        
        # Use ExcelFormatter to write data
        formatter = ExcelFormatter(self.template_manager)
        
        # Write headers
        formatter._write_headers(ws)
        
        # Write test cases
        for row_idx, test_case in enumerate(collection.test_cases, start=2):
            formatter._write_test_case(ws, row_idx, test_case, collection)
        
        # Apply styles
        formatter._apply_styles(ws)
        formatter._auto_adjust_columns(ws)
        
        self.sheets_created += 1
    
    def add_collections_by_module(self, collections: List[TestCaseCollection]):
        """Group collections by module and create one sheet per module.
        
        Args:
            collections: List of test case collections
        """
        # Group by module
        module_map = {}
        for collection in collections:
            # Get module from first test case (all should have same module)
            if collection.test_cases:
                module = getattr(collection.test_cases[0], 'module', 'General')
                if module not in module_map:
                    module_map[module] = []
                module_map[module].append(collection)
        
        # Create sheet for each module
        for module, module_collections in module_map.items():
            # Merge all collections for this module
            merged = self._merge_collections(module_collections)
            self.add_collection(merged, sheet_name=module)
    
    def add_collections_by_endpoint(self, collections: List[TestCaseCollection]):
        """Create one sheet per endpoint.
        
        Args:
            collections: List of test case collections
        """
        for collection in collections:
            self.add_collection(collection)
    
    def add_all_collections_single_sheet(self, collections: List[TestCaseCollection]):
        """Merge all collections into a single sheet.
        
        Args:
            collections: List of test case collections
        """
        merged = self._merge_all_collections(collections)
        self.add_collection(merged, sheet_name="All Test Cases")
    
    def save(self) -> bytes:
        """Save workbook as bytes.
        
        Returns:
            Excel file as bytes
        """
        buffer = BytesIO()
        self.workbook.save(buffer)
        return buffer.getvalue()
    
    def save_to_file(self, filepath: str):
        """Save workbook to file.
        
        Args:
            filepath: Path to save the Excel file
        """
        self.workbook.save(filepath)
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name to meet Excel requirements.
        
        Args:
            name: Original sheet name
            
        Returns:
            Sanitized sheet name (max 31 chars, no special chars)
        """
        # Remove invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Truncate to 31 characters (Excel limit)
        if len(name) > 31:
            name = name[:28] + "..."
        
        # Ensure uniqueness
        existing_names = [sheet.title for sheet in self.workbook.worksheets]
        original_name = name
        counter = 1
        while name in existing_names:
            suffix = f"_{counter}"
            max_length = 31 - len(suffix)
            name = original_name[:max_length] + suffix
            counter += 1
        
        return name
    
    def _generate_sheet_name(self, collection: TestCaseCollection) -> str:
        """Generate sheet name from collection.
        
        Args:
            collection: Test case collection
            
        Returns:
            Generated sheet name
        """
        # Use method and path
        name = f"{collection.method}_{collection.path.replace('/', '_')}"
        return self._sanitize_sheet_name(name)
    
    def _merge_collections(self, collections: List[TestCaseCollection]) -> TestCaseCollection:
        """Merge multiple collections into one.
        
        Args:
            collections: Collections to merge
            
        Returns:
            Merged collection
        """
        if not collections:
            return TestCaseCollection(
                endpoint_id="merged",
                method="MERGED",
                path="/merged",
                test_cases=[]
            )
        
        # Use first collection as base
        first = collections[0]
        all_test_cases = []
        
        # Collect all test cases
        for collection in collections:
            all_test_cases.extend(collection.test_cases)
        
        # Renumber test IDs
        for i, test_case in enumerate(all_test_cases, 1):
            test_case.test_id = i
        
        return TestCaseCollection(
            endpoint_id=f"merged_{first.endpoint_id}",
            method=first.method,
            path=first.path,
            summary=f"Merged: {first.summary or first.path}",
            description=f"Merged test cases from {len(collections)} endpoints",
            test_cases=all_test_cases,
            metadata=first.metadata
        )
    
    def _merge_all_collections(self, collections: List[TestCaseCollection]) -> TestCaseCollection:
        """Merge all collections into a single collection.
        
        Args:
            collections: All collections to merge
            
        Returns:
            Single merged collection
        """
        all_test_cases = []
        
        for collection in collections:
            all_test_cases.extend(collection.test_cases)
        
        # Renumber test IDs
        for i, test_case in enumerate(all_test_cases, 1):
            test_case.test_id = i
        
        return TestCaseCollection(
            endpoint_id="all",
            method="ALL",
            path="/all",
            summary="All Test Cases",
            description=f"Merged test cases from {len(collections)} endpoints",
            test_cases=all_test_cases
        )